import copy
import pandas
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import Alignment

from libs.misc import Utils

def read(filename, first_row_as_header=True):
  # Load the Excel file
  wb = load_workbook(filename)
  # Get the first sheet
  ws = wb.active
  # Get the data from the sheet
  data = ws.values
  # Get the header row
  if (first_row_as_header):
    columns = next(data)
    data = list(data)
  else:
    columns = None
  # Create the DataFrame
  df = pandas.DataFrame(data, columns=columns)
  # Convert the DataFrame to a dictionary and return it
  return df.to_dict(orient='records')

def fromCSV(csv_file, excel_file):
  df = pandas.read_csv(csv_file)
  df.to_excel(excel_file, index=False)

def fromJSON(data, filename):
  fdata = {}
  # Create copy of data to avoid modifying the original data
  p_data = copy.deepcopy(data)
  # Check if first level of data is a list, it means is a single tab data structure
  if (isinstance(p_data, list)):
    p_data = { 'Sheet1': p_data }
  # Loop through the data to create the final data structure
  for tab in p_data:
    # Loop through first tab row to get the column titles
    fdata[tab] = {}
    for row in p_data[tab]:
      for key in row:
        if (key not in fdata[tab]):
          fdata[tab][key] = []
    # Loop through the data to fill the columns
    for row in p_data[tab]:
      for key in row:
        value = row[key]
        fdata[tab][key].append(value)
  fromTabData(fdata, filename)

def fromTabData(data, filename):
  # Create a Pandas Excel writer using openpyxl as the engine
  writer = pandas.ExcelWriter(filename, engine='openpyxl')
  # Loops through data first level keys (tabs)
  for tab in data:
    # Loop through the tabs data to find cell having a list and value and convert it to a string
    for row_key in data[tab]:
      row_data = data[tab][row_key]
      for cell_index in range(len(row_data)):
        if (isinstance(row_data[cell_index], list)):
          row_data[cell_index] = '\n'.join(row_data[cell_index])
    # Create a DataFrame from the tab data
    df = pandas.DataFrame(data[tab])
    # Write the DataFrame to the Excel sheet
    df.to_excel(writer, sheet_name=tab, index=False)
  # Save the Excel file
  writer.close()
  # Format the Excel file
  _format(filename, max_column_width=20000)

def _format(filename, max_column_width=None):
  int_style = NamedStyle(name='englishNumberFormat', number_format='#,##0') # Define integer style
  float_style = NamedStyle(name='englishNumberFormat', number_format='#,##0.00') # Define float style
  percentage_style = NamedStyle(name='percentageFormat', number_format='0.00%')  # Define percentage style
  percentage_style_int = NamedStyle(name='percentageFormat', number_format='0%')  # Define percentage style
  # Load the workbook
  wb = load_workbook(filename)
  # Iterate over the worksheets
  for ws in wb.worksheets:
    # Set auto filter for the worksheet
    ws.auto_filter.ref = ws.dimensions
    # Iterate over the rows to format the cells
    for index, row in enumerate(ws.iter_rows()):
      for cell in row:
        if (index == 0):
          cell.alignment = Alignment(horizontal='center', vertical='center')
        elif ('\n' in str(cell.value)):
          cell.alignment = Alignment(wrap_text=True, vertical='center')
        else:
          cell.alignment = Alignment(vertical='center')
        # Text custom format
        _applyCustomFormat(cell)
        # Numbers format
        is_number = Utils.isNumber(cell.value, return_type=True)
        if (is_number == 'percentage'):
          #cell.value = cell.value.replace('%', '')
          if ('.' in cell.value or ',' in cell.value):
            cell.value = float(cell.value.replace('%', '')) / 100  # Convert string percentage to float
            cell.style = percentage_style  # Apply percentage style
          else:
            cell.value = int(cell.value.replace('%', '')) / 100
            cell.style = percentage_style_int  # Apply percentage style
        if (is_number == 'int'):
          cell.style = int_style
        elif (is_number == 'float'):
          cell.style = float_style
    # Set the width of each column
    for column in ws.columns:
      max_length = 0
      column = [cell for cell in column]
      cell_n = -1
      cell_longer = -1
      for cell in column:
        cell_n += 1
        try:
          value_parts = cell.value.split('\n')
          for value in value_parts:
            cell_length = len(str(value).strip())
            if (cell_length > max_length):
              max_length = cell_length
              cell_longer = cell_n
        except:
          pass
      if (cell_longer == 0):
        max_length += 5
      else:
        max_length += 2
      if (max_column_width and max_length > max_column_width):
        max_length = max_column_width
      ws.column_dimensions[column[0].column_letter].width = max_length
  # Save the workbook
  wb.save(filename)

def _applyCustomFormat(cell):
  font = None
  color = None
  value = str(cell.value)
  # Veryfies custom bold format
  if ('@bold@' in value):
    font = True
    value = value.replace('@bold@', '').strip()
  # Veryfies custom color format
  pattern = r"@color:(.*?)@"
  color_match = re.search(pattern, value)
  if color_match:
    color = color_match.group(1).replace('#', '')
    value = re.sub(pattern, '', value)
  # Applies custom format
  if (font and color):
    cell.font = Font(bold=True, color=color)
    cell.value = value.strip()
  elif (font):
    cell.font = Font(bold=True)
    cell.value = value.strip()
  elif (color):
    cell.font = Font(color=color)
    cell.value = value.strip()